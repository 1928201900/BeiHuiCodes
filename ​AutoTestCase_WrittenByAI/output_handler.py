from pathlib import Path
import openpyxl
from datetime import datetime
from typing import List, Dict, Any, Optional
import json

'''
定义了 OutputHandler 类，负责将生成的测试用例保存到 Excel 文件中。
根据新的格式要求，将测试用例按照特定的层级结构输出到 Excel 中。
'''

class OutputHandler:
    def __init__(self, config):
        self.config = config
        
    def save_to_excel(self, test_cases: List[Dict[str, Any]], 
                     output_path: Optional[Path] = None) -> Path:
        """保存测试用例到Excel文件"""
        if not test_cases:
            print("❌ 没有测试用例可保存")
            return None
            
        if not output_path:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = self.config.OUTPUTS_DIR / f"TestCases_{timestamp}.xlsx"
            
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # 表头
        headers = [
            "Object Type", "Name", "Short Description / Action", 
            "Expected Result", "input signal", "output signal"
        ]
        ws.append(headers)
        
        # 设置数据验证
        dv = openpyxl.worksheet.datavalidation.DataValidation(type="list", formula1='"Function,Feature,Test Group,Test Case,Precondition,Test Step"')
        ws.add_data_validation(dv)

        # 临时存储已添加的 Function、Feature 和 Test Group
        added_functions = set()
        added_features = set()
        added_test_groups = set()

        for case in test_cases:
            function_name = self._extract_function_name(case.get("description", ""))
            feature_name = self._extract_feature(case.get("coverage", []))
            test_group_name = self._extract_test_group(case.get("coverage", []))
            test_case_name = case.get("description", "")

            # 添加 Function 行
            if function_name not in added_functions:
                ws.append(["Function", function_name, "", "", "", "", "", "", "", ""])
                dv.add(f'A{ws.max_row}')
                added_functions.add(function_name)

            # 添加 Feature 行
            feature_key = (function_name, feature_name)
            if feature_key not in added_features:
                ws.append(["Feature", feature_name, "", "", "", "", "", "", "", ""])
                dv.add(f'A{ws.max_row}')
                added_features.add(feature_key)

            # 添加 Test Group 行
            test_group_key = (feature_name, test_group_name)
            if test_group_key not in added_test_groups:
                ws.append(["Test Group", test_group_name, "", "", "", "", "", "", "", ""])
                dv.add(f'A{ws.max_row}')
                added_test_groups.add(test_group_key)

            # 添加 Test Case 行
            ws.append(["Test Case", test_case_name, "", "", "", "", "", "", "", ""])
            dv.add(f'A{ws.max_row}')

            # 添加 Precondition 行
            preconditions = case.get("precondition", [])
            precondition_str = ',\n'.join([f"{i + 1}. {p}" for i, p in enumerate(preconditions)])
            input_signal_dict = case.get("input_signal", {})
            input_signal_str = ',\n'.join([f"{key}={value}" for key, value in input_signal_dict.items()])
            ws.append(["Precondition", "", precondition_str, "", input_signal_str, "", "", "", "", ""])
            dv.add(f'A{ws.max_row}')

            # 添加 Test Step 行
            steps = case.get("steps", [])
            expected_results = case.get("expected", [])
            output_signal = case.get("output_signal", "")
            for step in steps:
                ws.append([
                    "Test Step",
                    "",
                    step,
                    ', '.join(expected_results),
                    input_signal_str,
                    output_signal,
                    "",
                    "",
                    "",
                    ""
                ])
                dv.add(f'A{ws.max_row}')

        # 设置自动换行样式
        from openpyxl.styles import Alignment
        wrap_alignment = Alignment(wrap_text=True)
        
        # 为Short Description / Action列(C列)设置自动换行
        for row in ws.iter_rows(min_col=3, max_col=3):
            for cell in row:
                cell.alignment = wrap_alignment
                
        # 为input signal列(E列)设置自动换行
        for row in ws.iter_rows(min_col=5, max_col=5):
            for cell in row:
                cell.alignment = wrap_alignment

        # 统一设置所有列宽为30
        for col in 'ABCDEFGHIJ':  # 假设表格有10列
            ws.column_dimensions[col].width = 30


        # 保存文件
        wb.save(output_path)
        print(f"✅ 测试用例已保存至: {output_path}")
        return output_path
    
    def _extract_function_name(self, description: str) -> str:
        """从描述中提取功能名称"""
        if "倒车灯" in description:
            return "外灯控制"
        elif "门锁" in description:
            return "门锁控制"
        elif "雨刮" in description:
            return "雨刮控制"
        elif "电源" in description:
            return "电源管理"
        else:
            return "功能控制"
    
    def _extract_feature(self, coverage: List[str]) -> str:
        """从覆盖项中提取特性名称"""
        for item in coverage:
            if "倒车灯" in item:
                return "倒车灯功能"
            elif "门锁" in item:
                return "门锁功能"
            elif "雨刮" in item:
                return "雨刮功能"
            elif "电源" in item:
                return "电源管理功能"
        return "其他功能"
    
    def _extract_test_group(self, coverage: List[str]) -> str:
        """从覆盖项中提取测试组名称"""
        for item in coverage:
            if "倒车灯" in item:
                return "倒车灯"
            elif "门锁" in item:
                return "门锁"
            elif "雨刮" in item:
                return "雨刮"
            elif "电源" in item:
                return "电源"
        return "其他"