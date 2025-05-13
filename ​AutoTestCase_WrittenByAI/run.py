import os
import re
import json
import openpyxl
import PyPDF2
from datetime import datetime
from pathlib import Path
from openai import OpenAI
from typing import List, Dict, Any, Optional, Tuple

# ===== 配置区 =====
class Config:
    def __init__(self):
        self.BASE_DIR = Path(r"D:\codes\autotest")  # 项目根目录
        self.API_KEY = "sk-ad87abf37f884cd08455d87fe2dabbf8"  # 替换为您的阿里云API Key
        self.BASE_URL = "https://dashscope.aliyuncs.com/compatible-mode/v1"  # 阿里云API地址
        self.MODEL = "qwen-plus"  # 使用的模型
        
        # 路径配置
        self.INPUTS_DIR = self.BASE_DIR / "inputs"
        self.OUTPUTS_DIR = self.BASE_DIR / "outputs"
        self.CONFIG_PATH = self.BASE_DIR / "config"
        
        # 创建必要目录
        for dir_path in [self.OUTPUTS_DIR, self.INPUTS_DIR, self.CONFIG_PATH]:
            dir_path.mkdir(exist_ok=True)

# ===== 文档解析 =====
class DocumentParser:
    def __init__(self, config: Config):
        self.config = config
        self.signal_cache = {}
        
    def parse_pdf(self, pdf_path: Optional[Path] = None) -> List[Dict[str, Any]]:
        """解析功能规范PDF文件，返回结构化需求"""
        pdf_path = pdf_path or self.config.INPUTS_DIR / "功能规范-第七章.pdf"
        if not pdf_path.exists():
            raise FileNotFoundError(f"PDF文件不存在: {pdf_path}")
            
        with open(pdf_path, 'rb') as f:
            reader = PyPDF2.PdfReader(f)
            text = "\n".join([page.extract_text() for page in reader.pages])
        
        # 提取章节标题和内容
        sections = self._extract_sections(text)
        
        # 提取功能和信号相关信息
        requirements = []
        for section in sections:
            if "功能" in section["title"] or "需求" in section["title"] or "工作条件" in section["title"]:
                functions = self._extract_functions(section["content"])
                requirements.extend(functions)
                
            elif "信号" in section["title"] or "CAN" in section["title"]:
                signals = self._extract_signals(section["content"])
                self.signal_cache.update(signals)
                
        return requirements
    
    def parse_excel(self, excel_path: Optional[Path] = None) -> Dict[str, Dict[str, Any]]:
        """解析CAN信号矩阵Excel文件，返回信号字典"""
        excel_path = excel_path or self.config.INPUTS_DIR / "CAN信号矩阵-第七章.xlsx"
        if not excel_path.exists():
            raise FileNotFoundError(f"Excel文件不存在: {excel_path}")
            
        wb = openpyxl.load_workbook(excel_path, read_only=True)
        ws = wb.active
        
        signals = {}
        
        # 获取表头行
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [str(cell) for cell in header_row if cell is not None]
        
        # 映射常见表头变体到标准字段
        header_mapping = {
            "信号名称": ["信号名称", "Signal Name", "信号"],
            "消息名称": ["消息名称", "Message Name", "消息"],
            "起始位": ["起始位", "Start Bit", "位起始"],
            "位长度": ["位长度", "Bit Length", "长度"],
            "比例因子": ["比例因子", "Factor", "缩放因子"],
            "偏置": ["偏置", "Offset", "偏移"],
            "单位": ["单位", "Unit"],
            "最小值": ["最小值", "Min Value", "最小值"],
            "最大值": ["最大值", "Max Value", "最大值"]
        }
        
        # 查找每个标准字段对应的列索引
        header_indices = {}
        for std_field, alternatives in header_mapping.items():
            for idx, header in enumerate(headers):
                if any(alt.lower() in header.lower() for alt in alternatives):
                    header_indices[std_field] = idx
                    break
        
        # 解析数据行
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(cell is None for cell in row):
                continue
                
            signal_name = row[header_indices.get("信号名称", 0)]
            if not signal_name or not isinstance(signal_name, str):
                continue
                
            signal_info = {
                "message_name": row[header_indices.get("消息名称", 1)],
                "start_bit": row[header_indices.get("起始位", 2)],
                "bit_length": row[header_indices.get("位长度", 3)],
                "factor": row[header_indices.get("比例因子", 4)],
                "offset": row[header_indices.get("偏置", 5)],
                "unit": row[header_indices.get("单位", 6)],
                "value_range": f"{row[header_indices.get('最小值', 7)]}~{row[header_indices.get('最大值', 8)]}"
            }
            signals[signal_name] = signal_info
            
        # 合并PDF中提取的信号信息
        signals.update(self.signal_cache)
        return signals
    
    def _extract_sections(self, text: str) -> List[Dict[str, str]]:
        """将文本按章节分割"""
        # 匹配章节标题的模式
        section_patterns = [
            re.compile(r'第(\d+)\s*章\s*([^\n]+)'),  # 标准章节格式
            re.compile(r'(\d+)\s*、\s*([^\n]+)'),   # 数字加顿号格式
            re.compile(r'(\d+)\s*\.(\d+)\s*([^\n]+)'),  # 数字加小数点格式
        ]
        
        sections = []
        current_title = None
        current_content = []
        
        for line in text.split('\n'):
            line = line.strip()
            if not line:
                continue
                
            # 检查是否为新章节
            section_match = None
            for pattern in section_patterns:
                match = pattern.search(line)
                if match:
                    section_match = match
                    break
                    
            if section_match:
                if current_title:
                    sections.append({
                        "title": current_title,
                        "content": "\n".join(current_content)
                    })
                current_title = section_match.group(0)
                current_content = [line]
            else:
                current_content.append(line)
                
        # 添加最后一个章节
        if current_title and current_content:
            sections.append({
                "title": current_title,
                "content": "\n".join(current_content)
            })
            
        return sections
    
    def _extract_functions(self, content: str) -> List[Dict[str, Any]]:
        """从内容中提取功能需求"""
        functions = []
        
        # 尝试多种可能的功能定义模式
        patterns = [
            re.compile(r'(\d+\.\d+)\s*([^\n]+)'),  # 数字.数字 格式
            re.compile(r'(\d+)\s*、\s*([^\n]+)'),  # 数字、格式
            re.compile(r'([A-Z]\d+)\s*([^\n]+)'),  # 字母数字格式
        ]
        
        current_id = None
        current_desc = []
        
        for line in content.split('\n'):
            line = line.strip()
            if not line:
                continue
                
            # 检查是否为新功能
            func_match = None
            for pattern in patterns:
                match = pattern.search(line)
                if match:
                    func_match = match
                    break
                    
            if func_match:
                if current_id:
                    functions.append({
                        "id": current_id,
                        "description": "\n".join(current_desc),
                        "type": self._classify_function("\n".join(current_desc))
                    })
                current_id = func_match.group(1)
                current_desc = [func_match.group(2)]
            else:
                current_desc.append(line)
                
        # 添加最后一个功能
        if current_id and current_desc:
            functions.append({
                "id": current_id,
                "description": "\n".join(current_desc),
                "type": self._classify_function("\n".join(current_desc))
            })
            
        return functions
    
    def _extract_signals(self, content: str) -> Dict[str, Dict[str, Any]]:
        """从内容中提取信号信息"""
        signals = {}
        
        # 匹配信号定义的模式
        signal_patterns = [
            re.compile(r'([^\s]+)\s*信号\s*[:：]\s*([^\n]+)'),  # 信号名称:描述 格式
            re.compile(r'信号\s*[:：]\s*([^\s]+)\s*([^\n]+)'),  # 信号:名称 描述 格式
            re.compile(r'([^\s]+)\s*:\s*([^\n]+)'),  # 名称:描述 通用格式
        ]
        
        for line in content.split('\n'):
            line = line.strip()
            if not line:
                continue
                
            # 尝试匹配信号定义
            signal_match = None
            for pattern in signal_patterns:
                match = pattern.search(line)
                if match:
                    signal_match = match
                    break
                    
            if signal_match:
                signal_name = signal_match.group(1)
                signal_desc = signal_match.group(2)
                
                # 提取信号属性
                properties = {}
                for prop in re.findall(r'([^，,:：]+)[:：]([^，,:：]+)', signal_desc):
                    key, value = prop
                    properties[key.strip()] = value.strip()
                    
                signals[signal_name] = properties
                
        return signals
    
    def _classify_function(self, description: str) -> str:
        """根据描述对功能进行分类"""
        if "控制" in description or "调节" in description:
            return "Control"
        elif "监测" in description or "检测" in description:
            return "Monitoring"
        elif "保护" in description:
            return "Protection"
        elif "通信" in description:
            return "Communication"
        else:
            return "Other"

# ===== 测试用例生成 =====
class TestCaseGenerator:
    def __init__(self, config: Config):
        self.config = config
        self.client = OpenAI(
            api_key=config.API_KEY,
            base_url=config.BASE_URL
        )
        
    def generate(self, requirements: List[Dict[str, Any]], 
                 signals: Dict[str, Dict[str, Any]]) -> List[Dict[str, Any]]:
        """根据需求和信号生成测试用例"""
        if not requirements:
            print("⚠️ 警告：没有检测到功能需求，将使用示例测试用例")
            return self._generate_example_test_cases(signals)
            
        # 构建提示词
        prompt = self._build_prompt(requirements, signals)
        
        # 调用AI生成测试用例
        response = self._call_ai(prompt)
        if not response:
            return []
            
        # 解析AI响应
        try:
            test_cases = self._parse_response(response, requirements, signals)
            return test_cases
        except Exception as e:
            print(f"⚠️ 解析AI响应失败: {str(e)}")
            print(f"原始响应内容:\n{response[:500]}...")
            return []
    
    def _build_prompt(self, requirements: List[Dict[str, Any]], 
                     signals: Dict[str, Dict[str, Any]]) -> str:
        """构建AI生成测试用例的提示词"""
        # 示例测试用例格式说明
        example_format = """
【测试用例格式】[
  {
    "description": "挂R档，倒车灯点亮",
    "coverage": ["需求1.1", "VCU_ActGear"],
    "input_signal": {
      "IGN1_RELAY_FB": "ON",
      "VCU_ActGear": "0x9",
      "VCU_ActGear_VD": "0x0"
    },
    "output_signal": "倒车灯点亮",
    "precondition": [
      "车辆处于ON电源模式",
      "当前档位为P档",
      "档位信号有效"
    ],
    "steps": [
      "1. 确认车辆处于ON电源模式",
      "2. 确认当前档位为P档",
      "3. 切换档位至R档",
      "4. 观察倒车灯状态"
    ],
    "expected": [
      "倒车灯应点亮"
    ]
  }
]        """
        
        # 构建提示词
        prompt = f"""
你是一位专业的汽车电子测试工程师，擅长根据功能规范和CAN信号矩阵生成全面的测试用例。

【功能规范】
{json.dumps(requirements, indent=2, ensure_ascii=False)}

【CAN信号矩阵】
{json.dumps(signals, indent=2, ensure_ascii=False)}

请基于以上信息，生成符合以下要求的测试用例：
1. 测试用例描述简洁，直接说明测试场景和验证内容，如“挂R档，倒车灯点亮”。
2. 覆盖所有功能需求，包括正常、异常和边界情况
3. 每个测试用例必须包含：
   - 测试描述（简明扼要地说明测试内容）
   - 覆盖的需求ID/信号ID
   - 详细测试步骤（使用序号开头，如"1. 操作内容"）
   - 预期结果（明确的预期行为）
   - 输入信号（使用信号名称和有效值）
   - 输出信号（预期的信号变化）
   - 前置条件（执行测试前必须满足的条件）
4. 输出格式：[
  {{
    "description": "测试描述",
    "coverage": ["需求ID/信号ID"],
    "input_signal": {{
      "信号名称": "信号值"
    }},
    "output_signal": "输出信号变化",
    "precondition": ["前置条件1", "前置条件2"],
    "steps": ["步骤1", "步骤2"],
    "expected": ["预期结果1", "预期结果2"]
  }}
]
以下是一个测试用例的示例格式：
{example_format}
        """
        
        return prompt
    
    def _call_ai(self, prompt: str) -> Optional[str]:
        """调用AI生成测试用例"""
        try:
            print("正在调用AI生成测试用例...")
            completion = self.client.chat.completions.create(
                model=self.config.MODEL,
                messages=[
                    {"role": "system", "content": "你是一个专业的汽车电子测试工程师，擅长根据功能规范和CAN信号矩阵生成全面的测试用例。"},
                    {"role": "user", "content": prompt}
                ],
                temperature=0.3,  # 降低随机性，提高确定性
                max_tokens=4000
            )
            return completion.choices[0].message.content
        except Exception as e:
            print(f"⚠️ API调用失败: {str(e)}")
            return None
    
    def _parse_response(self, response: str, 
                       requirements: List[Dict[str, Any]], 
                       signals: Dict[str, Dict[str, Any]]) -> List[Dict[str, Any]]:
        """解析AI响应并转换为内部格式"""
        # 提取JSON部分
        try:
            # 尝试多种JSON解析方式
            start_idx = response.find('[')
            end_idx = response.rfind(']') + 1
            if start_idx < 0 or end_idx <= start_idx:
                # 尝试作为对象解析
                start_idx = response.find('{')
                end_idx = response.rfind('}') + 1
                if start_idx >= 0 and end_idx > start_idx:
                    json_str = response[start_idx:end_idx]
                    data = json.loads(json_str)
                    if isinstance(data, dict) and "cases" in data:
                        return data["cases"]
                    else:
                        return [data] if isinstance(data, dict) else []
                else:
                    raise ValueError("无法找到有效的JSON结构")
            else:
                json_str = response[start_idx:end_idx]
                return json.loads(json_str)
                
        except Exception as e:
            print(f"⚠️ 解析AI响应失败: {str(e)}")
            print(f"原始响应内容:\n{response[:500]}...")
            return []
    
    def _generate_example_test_cases(self, signals: Dict[str, Dict[str, Any]]) -> List[Dict[str, Any]]:
        """生成示例测试用例（当没有检测到需求时使用）"""
        example_cases = [
            {
                "description": "挂R档，倒车灯点亮",
                "coverage": ["示例需求", "VCU_ActGear"],
                "input_signal": {
                    "IGN1_RELAY_FB": "ON",
                    "VCU_ActGear": "0x9",
                    "VCU_ActGear_VD": "0x0"
                },
                "output_signal": "倒车灯点亮",
                "precondition": [
                    "车辆处于ON电源模式",
                    "当前档位为P档",
                    "档位信号有效"
                ],
                "steps": [
                    "1. 确认车辆处于ON电源模式",
                    "2. 确认当前档位为P档",
                    "3. 切换档位至R档",
                    "4. 观察倒车灯状态"
                ],
                "expected": [
                    "倒车灯应点亮"
                ]
            },
            {
                "description": "下OFF电，倒车灯熄灭",
                "coverage": ["示例需求", "IGN1_RELAY_FB"],
                "input_signal": {
                    "IGN1_RELAY_FB": "OFF",
                    "VCU_ActGear": "0x9",
                    "VCU_ActGear_VD": "0x0"
                },
                "output_signal": "倒车灯熄灭",
                "precondition": [
                    "车辆处于ON电源模式",
                    "当前档位为R档",
                    "倒车灯处于点亮状态"
                ],
                "steps": [
                    "1. 确认车辆处于ON电源模式",
                    "2. 确认当前档位为R档且倒车灯点亮",
                    "3. 将车辆切换至OFF电源模式",
                    "4. 观察倒车灯状态"
                ],
                "expected": [
                    "倒车灯应熄灭"
                ]
            },
            {
                "description": "挂N档，倒车灯熄灭",
                "coverage": ["示例需求", "VCU_ActGear"],
                "input_signal": {
                    "IGN1_RELAY_FB": "ON",
                    "VCU_ActGear": "0xA",
                    "VCU_ActGear_VD": "0x0"
                },
                "output_signal": "倒车灯熄灭",
                "precondition": [
                    "车辆处于ON电源模式",
                    "当前档位为R档",
                    "倒车灯处于点亮状态"
                ],
                "steps": [
                    "1. 确认车辆处于ON电源模式",
                    "2. 确认当前档位为R档且倒车灯点亮",
                    "3. 切换档位至N档",
                    "4. 观察倒车灯状态"
                ],
                "expected": [
                    "倒车灯应熄灭"
                ]
            },
            {
                "description": "挂D档，倒车灯熄灭",
                "coverage": ["示例需求", "VCU_ActGear"],
                "input_signal": {
                    "IGN1_RELAY_FB": "ON",
                    "VCU_ActGear": "0x1",
                    "VCU_ActGear_VD": "0x0"
                },
                "output_signal": "倒车灯熄灭",
                "precondition": [
                    "车辆处于ON电源模式",
                    "当前档位为R档",
                    "倒车灯处于点亮状态"
                ],
                "steps": [
                    "1. 确认车辆处于ON电源模式",
                    "2. 确认当前档位为R档且倒车灯点亮",
                    "3. 切换档位至D档",
                    "4. 观察倒车灯状态"
                ],
                "expected": [
                    "倒车灯应熄灭"
                ]
            },
            {
                "description": "挂P档，倒车灯熄灭",
                "coverage": ["示例需求", "VCU_ActGear"],
                "input_signal": {
                    "IGN1_RELAY_FB": "ON",
                    "VCU_ActGear": "0xB",
                    "VCU_ActGear_VD": "0x0"
                },
                "output_signal": "倒车灯熄灭",
                "precondition": [
                    "车辆处于ON电源模式",
                    "当前档位为R档",
                    "倒车灯处于点亮状态"
                ],
                "steps": [
                    "1. 确认车辆处于ON电源模式",
                    "2. 确认当前档位为R档且倒车灯点亮",
                    "3. 切换档位至P档",
                    "4. 观察倒车灯状态"
                ],
                "expected": [
                    "倒车灯应熄灭"
                ]
            },
            {
                "description": "挡位无效，挂R档，倒车灯不点亮",
                "coverage": ["示例需求", "VCU_ActGear_VD"],
                "input_signal": {
                    "IGN1_RELAY_FB": "ON",
                    "VCU_ActGear": "0x9",
                    "VCU_ActGear_VD": "0x1"
                },
                "output_signal": "倒车灯不点亮",
                "precondition": [
                    "车辆处于ON电源模式",
                    "当前档位为P档",
                    "档位信号无效"
                ],
                "steps": [
                    "1. 确认车辆处于ON电源模式",
                    "2. 确认当前档位为P档",
                    "3. 模拟档位信号无效状态",
                    "4. 切换档位至R档",
                    "5. 观察倒车灯状态"
                ],
                "expected": [
                    "倒车灯不应点亮"
                ]
            }
        ]
        return example_cases

# ===== 输出处理器 =====
class OutputHandler:
    def __init__(self, config: Config):
        self.config = config
        
    def save_to_excel(self, test_cases: List[Dict[str, Any]], 
                     output_path: Optional[Path] = None) -> Path:
        """保存测试用例到Excel文件"""
        if not test_cases:
            print("❌ 没有测试用例可保存")
            return None
            
        if not output_path:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = self.config.OUTPUTS_DIR / f"TestCases_{timestamp}.xlsx"
            
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # 表头
        headers = [
            "Object Type", "Name", "Short Description / Action", 
            "Expected Result", "input signal", "output signal",
            "Feature", "Test Group", "Test Case", "Precondition"
        ]
        ws.append(headers)
        
        # 填充数据
        for case in test_cases:
            # 转换为标准格式
            function_row = {
                "Object Type": "Function",
                "Name": self._extract_function_name(case.get("description", "")),
                "Short Description / Action": case.get("description", ""),
                "Expected Result": ", ".join(case.get("expected", [])),
                "input signal": json.dumps(case.get("input_signal", {}), ensure_ascii=False),
                "output signal": case.get("output_signal", ""),
                "Feature": self._extract_feature(case.get("coverage", [])),
                "Test Group": self._extract_test_group(case.get("coverage", [])),
                "Test Case": case.get("description", ""),
                "Precondition": "\n".join(case.get("precondition", []))
            }
            
            # 添加功能行
            ws.append([
                function_row["Object Type"],
                function_row["Name"],
                function_row["Short Description / Action"],
                function_row["Expected Result"],
                function_row["input signal"],
                function_row["output signal"],
                function_row["Feature"],
                function_row["Test Group"],
                function_row["Test Case"],
                function_row["Precondition"]
            ])
            
            # 添加测试步骤行
            for step in case.get("steps", []):
                ws.append([
                    "Test Step",
                    function_row["Name"],
                    step,
                    "",
                    "",
                    "",
                    function_row["Feature"],
                    function_row["Test Group"],
                    function_row["Test Case"],
                    ""
                ])
            
            # 添加空行分隔不同测试用例
            ws.append([])
            
        # 保存文件
        wb.save(output_path)
        print(f"✅ 测试用例已保存至: {output_path}")
        return output_path
    
    def _extract_function_name(self, description: str) -> str:
        """从描述中提取功能名称"""
        if "倒车灯" in description:
            return "外灯控制"
        elif "门锁" in description:
            return "门锁控制"
        elif "雨刮" in description:
            return "雨刮控制"
        elif "电源" in description:
            return "电源管理"
        else:
            return "功能控制"
    
    def _extract_feature(self, coverage: List[str]) -> str:
        """从覆盖项中提取特性名称"""
        for item in coverage:
            if "倒车灯" in item:
                return "倒车灯功能"
            elif "门锁" in item:
                return "门锁功能"
            elif "雨刮" in item:
                return "雨刮功能"
            elif "电源" in item:
                return "电源管理功能"
        return "其他功能"
    
    def _extract_test_group(self, coverage: List[str]) -> str:
        """从覆盖项中提取测试组名称"""
        for item in coverage:
            if "倒车灯" in item:
                return "倒车灯"
            elif "门锁" in item:
                return "门锁"
            elif "雨刮" in item:
                return "雨刮"
            elif "电源" in item:
                return "电源"
        return "其他"

# ===== 主流程 =====
def main():
    print("🚀 汽车电子测试用例生成系统 v2.0 (通用框架)")
    config = Config()
    
    try:
        # 1. 解析文档
        print("🔧 正在解析输入文档...")
        parser = DocumentParser(config)
        requirements = parser.parse_pdf()
        signals = parser.parse_excel()
        
        print(f"📑 识别到 {len(requirements)} 条功能需求")
        print(f"📶 识别到 {len(signals)} 个CAN信号")
        
        # 2. 生成测试用例
        print("⚡ 正在智能生成测试用例...")
        generator = TestCaseGenerator(config)
        test_cases = generator.generate(requirements, signals)
        
        if not test_cases:
            print("❌ 未能生成任何测试用例")
            return
            
        print(f"✅ 成功生成 {len(test_cases)} 个测试用例")
        
        # 3. 保存结果
        output_handler = OutputHandler(config)
        output_path = output_handler.save_to_excel(test_cases)
        
        if output_path:
            print(f"\n📊 生成统计:")
            print(f"- 生成测试用例数量: {len(test_cases)}")
            print(f"- 输出文件: {output_path}")
        
    except Exception as e:
        print(f"❌ 执行过程中发生错误: {str(e)}")
        import traceback
        traceback.print_exc()  # 打印详细的错误堆栈信息

if __name__ == "__main__":
    main()    