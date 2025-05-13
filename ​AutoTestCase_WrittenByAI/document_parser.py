from pathlib import Path
import re
import PyPDF2
import openpyxl
from typing import List, Dict, Any, Optional

'''
定义了 DocumentParser 类，用于解析功能规范 PDF 文件和 CAN 信号矩阵 Excel 文件。
它可以从 PDF 文件中提取结构化的功能需求，
从 Excel 文件中提取 CAN 信号信息，并将这些信息合并返回。
'''

class DocumentParser:
    def __init__(self, config):
        self.config = config
        self.signal_cache = {}
        
    def parse_pdf(self, pdf_path: Optional[Path] = None) -> List[Dict[str, Any]]:
        """解析功能规范PDF文件，返回结构化需求"""
        pdf_path = pdf_path or self.config.INPUTS_DIR / "功能规范-第七章.pdf"
        if not pdf_path.exists():
            raise FileNotFoundError(f"PDF文件不存在: {pdf_path}")
            
        with open(pdf_path, 'rb') as f:
            reader = PyPDF2.PdfReader(f)
            # 增强文本提取，处理可能的编码问题
            text = ""
            for page in reader.pages:
                try:
                    page_text = page.extract_text()
                    if page_text:
                        text += page_text + "\n"
                except Exception as e:
                    print(f"⚠️ 页面解析异常: {str(e)}")
                    continue
        
        # 增强章节分割逻辑
        sections = self._extract_sections(text)
        
        # 增强功能需求提取
        requirements = []
        for section in sections:
            # 扩展关键词识别
            keywords = ["功能", "需求", "工作条件", "要求", "规范", "specification"]
            if any(kw in section["title"] for kw in keywords):
                functions = self._extract_functions(section["content"])
                requirements.extend(functions)
                
            elif "信号" in section["title"] or "CAN" in section["title"]:
                signals = self._extract_signals(section["content"])
                self.signal_cache.update(signals)
                
        return requirements
    
    def parse_excel(self, excel_path: Optional[Path] = None) -> Dict[str, Dict[str, Any]]:
        """解析CAN信号矩阵Excel文件，返回信号字典"""
        excel_path = excel_path or self.config.INPUTS_DIR / "CAN信号矩阵-第七章.xlsx"
        if not excel_path.exists():
            raise FileNotFoundError(f"Excel文件不存在: {excel_path}")
            
        wb = openpyxl.load_workbook(excel_path, read_only=True)
        ws = wb.active
        
        signals = {}
        
        # 获取表头行
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [str(cell) for cell in header_row if cell is not None]
        
        # 映射常见表头变体到标准字段
        header_mapping = {
            "信号名称": ["信号名称", "Signal Name", "信号"],
            "消息名称": ["消息名称", "Message Name", "消息"],
            "起始位": ["起始位", "Start Bit", "位起始"],
            "位长度": ["位长度", "Bit Length", "长度"],
            "比例因子": ["比例因子", "Factor", "缩放因子"],
            "偏置": ["偏置", "Offset", "偏移"],
            "单位": ["单位", "Unit"],
            "最小值": ["最小值", "Min Value", "最小值"],
            "最大值": ["最大值", "Max Value", "最大值"]
        }
        
        # 查找每个标准字段对应的列索引
        header_indices = {}
        for std_field, alternatives in header_mapping.items():
            for idx, header in enumerate(headers):
                if any(alt.lower() in header.lower() for alt in alternatives):
                    header_indices[std_field] = idx
                    break
        
        # 解析数据行
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(cell is None for cell in row):
                continue
                
            signal_name = row[header_indices.get("信号名称", 0)]
            if not signal_name or not isinstance(signal_name, str):
                continue
                
            signal_info = {
                "message_name": row[header_indices.get("消息名称", 1)],
                "start_bit": row[header_indices.get("起始位", 2)],
                "bit_length": row[header_indices.get("位长度", 3)],
                "factor": row[header_indices.get("比例因子", 4)],
                "offset": row[header_indices.get("偏置", 5)],
                "unit": row[header_indices.get("单位", 6)],
                "value_range": f"{row[header_indices.get('最小值', 7)]}~{row[header_indices.get('最大值', 8)]}"
            }
            signals[signal_name] = signal_info
            
        # 合并PDF中提取的信号信息
        signals.update(self.signal_cache)
        return signals
    
    def _extract_sections(self, text: str) -> List[Dict[str, str]]:
        """将文本按章节分割"""
        # 匹配章节标题的模式
        section_patterns = [
            re.compile(r'第(\d+)\s*章\s*([^\n]+)'),  # 标准章节格式
            re.compile(r'(\d+)\s*、\s*([^\n]+)'),   # 数字加顿号格式
            re.compile(r'(\d+)\s*\.(\d+)\s*([^\n]+)'),  # 数字加小数点格式
        ]
        
        sections = []
        current_title = None
        current_content = []
        
        for line in text.split('\n'):
            line = line.strip()
            if not line:
                continue
                
            # 检查是否为新章节
            section_match = None
            for pattern in section_patterns:
                match = pattern.search(line)
                if match:
                    section_match = match
                    break
                    
            if section_match:
                if current_title:
                    sections.append({
                        "title": current_title,
                        "content": "\n".join(current_content)
                    })
                current_title = section_match.group(0)
                current_content = [line]
            else:
                current_content.append(line)
                
        # 添加最后一个章节
        if current_title and current_content:
            sections.append({
                "title": current_title,
                "content": "\n".join(current_content)
            })
            
        return sections
    
    def _extract_functions(self, content: str) -> List[Dict[str, Any]]:
        """从内容中提取功能需求"""
        functions = []
        
        # 增强模式匹配
        patterns = [
            re.compile(r'(\d+\.\d+(?:\.\d+)*)\s*([^\n]+)'),  # 支持多级编号
            re.compile(r'([A-Z]{2,3}_\d+)\s*([^\n]+)'),  # 支持类似"ECU_001"的编号
            re.compile(r'(\d+)\s*、\s*([^\n]+)'),  
            re.compile(r'([A-Z]\d+)\s*([^\n]+)'),
            re.compile(r'(?:需求|功能)\s*[:：]\s*([^\n]+)')  # 支持"需求："开头的行
        ]
        
        current_id = None
        current_desc = []
        
        for line in content.split('\n'):
            line = line.strip()
            if not line:
                continue
                
            # 检查是否为新功能
            func_match = None
            for pattern in patterns:
                match = pattern.search(line)
                if match:
                    func_match = match
                    break
                    
            if func_match:
                if current_id:
                    functions.append({
                        "id": current_id,
                        "description": "\n".join(current_desc),
                        "type": self._classify_function("\n".join(current_desc))
                    })
                current_id = func_match.group(1)
                current_desc = [func_match.group(2)] if len(func_match.groups()) > 1 else [line]
            else:
                current_desc.append(line)
                
        # 添加最后一个功能
        if current_id and current_desc:
            functions.append({
                "id": current_id,
                "description": "\n".join(current_desc),
                "type": self._classify_function("\n".join(current_desc))
            })
            
        return functions
    
    def _extract_signals(self, content: str) -> Dict[str, Dict[str, Any]]:
        """从内容中提取信号信息"""
        signals = {}
        
        # 匹配信号定义的模式
        signal_patterns = [
            re.compile(r'([^\s]+)\s*信号\s*[:：]\s*([^\n]+)'),  # 信号名称:描述 格式
            re.compile(r'信号\s*[:：]\s*([^\s]+)\s*([^\n]+)'),  # 信号:名称 描述 格式
            re.compile(r'([^\s]+)\s*:\s*([^\n]+)'),  # 名称:描述 通用格式
        ]
        
        for line in content.split('\n'):
            line = line.strip()
            if not line:
                continue
                
            # 尝试匹配信号定义
            signal_match = None
            for pattern in signal_patterns:
                match = pattern.search(line)
                if match:
                    signal_match = match
                    break
                    
            if signal_match:
                signal_name = signal_match.group(1)
                signal_desc = signal_match.group(2)
                
                # 提取信号属性
                properties = {}
                for prop in re.findall(r'([^，,:：]+)[:：]([^，,:：]+)', signal_desc):
                    key, value = prop
                    properties[key.strip()] = value.strip()
                    
                signals[signal_name] = properties
                
        return signals
    
    def _classify_function(self, description: str) -> str:
        """根据描述对功能进行分类"""
        if "控制" in description or "调节" in description:
            return "Control"
        elif "监测" in description or "检测" in description:
            return "Monitoring"
        elif "保护" in description:
            return "Protection"
        elif "通信" in description:
            return "Communication"
        else:
            return "Other"